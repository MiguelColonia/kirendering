"""
Tests de integración para los tres exportadores (IFC, DXF, XLSX).

Verifican que los archivos se generan, tienen contenido y son parseables.
El pipeline completo es: solver → builder → export_to_*.
"""

from pathlib import Path

import pytest

from cimiento.bim import export_to_dxf, export_to_ifc, export_to_xlsx, validate_ifc
from cimiento.geometry import build_building_from_solution
from cimiento.schemas import Program, Solar, Typology, TypologyMix

pytest_plugins = ["tests.fixtures.valid_cases"]


@pytest.fixture
def shared_building(sample_solar_rectangular: Solar, sample_typology_t2: Typology):
    from cimiento.solver import solve

    program = Program(
        project_id="exports-test",
        num_floors=1,
        floor_height_m=3.0,
        typologies=[sample_typology_t2],
        mix=[TypologyMix(typology_id="T2", count=5)],
    )
    solution = solve(sample_solar_rectangular, program)
    building = build_building_from_solution(sample_solar_rectangular, program, solution)
    return building, program


class TestDxfExport:
    def test_dxf_file_created(self, shared_building, tmp_path: Path) -> None:
        building, _ = shared_building
        path = tmp_path / "test.dxf"
        export_to_dxf(building, path)
        assert path.exists() and path.stat().st_size > 0

    def test_dxf_is_parseable(self, shared_building, tmp_path: Path) -> None:
        """El DXF debe ser abrirable con ezdxf."""
        import ezdxf

        building, _ = shared_building
        path = tmp_path / "test.dxf"
        export_to_dxf(building, path)
        doc = ezdxf.readfile(str(path))
        assert doc is not None

    def test_dxf_has_layers(self, shared_building, tmp_path: Path) -> None:
        """Deben existir las capas MUROS_EXT, MUROS_INT, PUERTAS y TEXTOS."""
        import ezdxf

        building, _ = shared_building
        path = tmp_path / "test.dxf"
        export_to_dxf(building, path)
        doc = ezdxf.readfile(str(path))
        layer_names = {layer.dxf.name for layer in doc.layers}
        for expected in ("MUROS_EXT", "MUROS_INT", "PUERTAS", "TEXTOS"):
            assert expected in layer_names, f"Capa {expected} no encontrada"

    def test_dxf_has_blocks_per_storey(self, shared_building, tmp_path: Path) -> None:
        """Debe haber un bloque DXF por cada planta del building."""
        import ezdxf

        building, _ = shared_building
        path = tmp_path / "test.dxf"
        export_to_dxf(building, path)
        doc = ezdxf.readfile(str(path))
        block_names = {b.name for b in doc.blocks}
        for storey in building.storeys:
            expected = f"PLANTA_{storey.level}"
            assert expected in block_names, f"Bloque {expected} no encontrado"


class TestXlsxExport:
    def test_xlsx_file_created(self, shared_building, tmp_path: Path) -> None:
        building, program = shared_building
        path = tmp_path / "test.xlsx"
        export_to_xlsx(building, program, path)
        assert path.exists() and path.stat().st_size > 0

    def test_xlsx_has_four_sheets(self, shared_building, tmp_path: Path) -> None:
        """El workbook debe tener exactamente 4 hojas."""
        from openpyxl import load_workbook

        building, program = shared_building
        path = tmp_path / "test.xlsx"
        export_to_xlsx(building, program, path)
        wb = load_workbook(str(path))
        assert set(wb.sheetnames) == {
            "Resumen", "Unidades", "Superficies por tipo", "Parámetros urbanísticos"
        }

    def test_xlsx_units_sheet_has_five_rows(self, shared_building, tmp_path: Path) -> None:
        """La hoja Unidades debe tener una fila por espacio (5 unidades)."""
        from openpyxl import load_workbook

        building, program = shared_building
        path = tmp_path / "test.xlsx"
        export_to_xlsx(building, program, path)
        wb = load_workbook(str(path))
        ws = wb["Unidades"]
        # Fila 1 = header; filas 2..6 = datos
        data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c for c in r)]
        assert len(data_rows) >= 5, f"Se esperaban ≥5 filas de datos, hay {len(data_rows)}"

    def test_xlsx_summary_contains_project_id(self, shared_building, tmp_path: Path) -> None:
        """La hoja Resumen debe incluir el project_id del programa."""
        from openpyxl import load_workbook

        building, program = shared_building
        path = tmp_path / "test.xlsx"
        export_to_xlsx(building, program, path)
        wb = load_workbook(str(path))
        ws = wb["Resumen"]
        cell_values = [c for row in ws.iter_rows(values_only=True) for c in row if c is not None]
        assert program.project_id in cell_values, "project_id no encontrado en la hoja Resumen"
