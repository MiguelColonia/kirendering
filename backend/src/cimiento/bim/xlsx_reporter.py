"""
Generador de informes XLSX para la jerarquía BIM de Cimiento.

Produce un workbook de openpyxl con cuatro hojas:

  "Resumen"               — Indicadores clave del proyecto.
  "Unidades"              — Tabla detallada de cada unidad residencial.
  "Superficies por tipo"  — Desglose por RoomType con totales.
  "Parámetros urbanísticos" — Edificabilidad, ocupación y número de plantas.

Todas las tablas llevan headers en negrita, bordes de celda y formato numérico
de miles con dos decimales.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from cimiento.schemas.architectural import Building, ParkingSpace, Space
from cimiento.schemas.program import Program
from cimiento.schemas.typology import RoomType

# ---------------------------------------------------------------------------
# Estilos reutilizables
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
_THIN_SIDE = Side(style="thin")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)
_NUM_FORMAT = "#,##0.00"
_INT_FORMAT = "#,##0"


# ---------------------------------------------------------------------------
# Función pública
# ---------------------------------------------------------------------------


def export_to_xlsx(building: Building, program: Program, output_path: Path) -> None:
    """
    Genera un workbook XLSX con métricas del building y lo escribe en output_path.

    :param building: Jerarquía BIM completa.
    :param program: Programa de necesidades asociado al building.
    :param output_path: Ruta del archivo .xlsx a crear.
    """
    wb = Workbook()
    wb.remove(wb.active)  # eliminar hoja vacía por defecto

    _sheet_summary(wb, building, program)
    _sheet_units(wb, building)
    _sheet_area_by_type(wb, building)
    _sheet_urban_params(wb, building, program)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


# ---------------------------------------------------------------------------
# Hojas
# ---------------------------------------------------------------------------


def _sheet_summary(wb: Workbook, building: Building, program: Program) -> None:
    ws = wb.create_sheet("Resumen")
    spaces = _all_spaces(building)
    residential_spaces = _residential_spaces(building)
    parking_spaces = _parking_spaces(building)
    total_area = sum(s.net_area_m2 for s in spaces)

    # Mix tipológico
    mix_str = ", ".join(
        f"{m.count}×{m.typology_id}" for m in program.mix
    ) if program.mix else "—"

    rows = [
        ("Proyecto", building.project_id),
        ("Edificio", building.name),
        ("Solar", building.solar_id),
        ("Número de plantas", len(building.storeys)),
        ("Total espacios", len(spaces)),
        ("Unidades residenciales", len(residential_spaces)),
        ("Plazas de aparcamiento", len(parking_spaces)),
        ("Área neta total (m²)", total_area),
        ("Mix tipológico", mix_str),
    ]

    for r, (label, value) in enumerate(rows, start=1):
        lbl_cell = ws.cell(row=r, column=1, value=label)
        val_cell = ws.cell(row=r, column=2, value=value)
        lbl_cell.font = _HEADER_FONT
        if isinstance(value, float):
            val_cell.number_format = _NUM_FORMAT
        _set_border(lbl_cell)
        _set_border(val_cell)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30


def _sheet_units(wb: Workbook, building: Building) -> None:
    ws = wb.create_sheet("Unidades")
    headers = [
        "ID Espacio", "Tipología", "Planta", "Área neta (m²)",
        "Área construida (m²)", "Ratio neta/construida",
    ]
    _write_header_row(ws, headers, row=1)

    for row_idx, space in enumerate(_residential_spaces(building), start=2):
        # Área construida estimada: área neta + 15 % de muros (estimación fase anteproyecto)
        gross = space.net_area_m2 * 1.15
        ratio = space.net_area_m2 / gross if gross > 0 else 0.0

        values = [
            space.id,
            space.typology_id or "—",
            space.floor_level,
            space.net_area_m2,
            gross,
            ratio,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _set_border(cell)
            if isinstance(val, float):
                if col_idx == 6:
                    cell.number_format = "0.00%"
                else:
                    cell.number_format = _NUM_FORMAT

    _autofit_columns(ws)


def _sheet_area_by_type(wb: Workbook, building: Building) -> None:
    ws = wb.create_sheet("Superficies por tipo")
    headers = ["Tipo de espacio", "Nº espacios", "Área total (m²)", "% sobre total"]
    _write_header_row(ws, headers, row=1)

    spaces = _all_spaces(building)
    total = sum(s.net_area_m2 for s in spaces)
    by_type: dict[str, list[float]] = {}
    for s in spaces:
        by_type.setdefault(s.room_type, []).append(s.net_area_m2)

    for row_idx, room_type in enumerate(RoomType, start=2):
        areas = by_type.get(room_type, [])
        subtotal = sum(areas)
        pct = subtotal / total if total > 0 else 0.0
        vals = [room_type.value, len(areas), subtotal, pct]
        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _set_border(cell)
            if col_idx == 3:
                cell.number_format = _NUM_FORMAT
            elif col_idx == 4:
                cell.number_format = "0.00%"

    # Fila de totales
    total_row = len(RoomType) + 2
    for col_idx, val in enumerate(["TOTAL", len(spaces), total, 1.0], start=1):
        cell = ws.cell(row=total_row, column=col_idx, value=val)
        cell.font = _HEADER_FONT
        _set_border(cell)
        if col_idx == 3:
            cell.number_format = _NUM_FORMAT
        elif col_idx == 4:
            cell.number_format = "0.00%"

    _autofit_columns(ws)


def _sheet_urban_params(wb: Workbook, building: Building, program: Program) -> None:
    """
    Hoja de parámetros urbanísticos básicos.

    Los valores de edificabilidad y ocupación se calculan de forma aproximada
    a partir de los datos disponibles en fase de anteproyecto. La superficie
    de solar no está disponible en este schema, así que se calcula a partir del
    área de los forjados de planta baja (aproximación conservadora).
    """
    ws = wb.create_sheet("Parámetros urbanísticos")

    spaces = _all_spaces(building)
    residential_spaces = _residential_spaces(building)
    parking_spaces = _parking_spaces(building)
    total_net = sum(s.net_area_m2 for s in spaces)
    num_storeys = len(building.storeys)
    total_gross = total_net * 1.15  # estimación con coeficiente de muros

    # Se usa área construida por planta como proxy (sin solar georreferenciado disponible)
    area_per_floor = total_gross / num_storeys if num_storeys > 0 else 0.0

    rows = [
        ("Número de plantas", num_storeys),
        ("Área neta total (m²)", total_net),
        ("Área construida total (m²)", total_gross),
        ("Área construida por planta (m²)", area_per_floor),
        ("Nº unidades residenciales", len(residential_spaces)),
        ("Nº plazas de aparcamiento", len(parking_spaces)),
        ("Altura total estimada (m)", sum(s.height_m for s in building.storeys)),
        ("Nota", "Valores orientativos de anteproyecto. Revisar con solar georreferenciado."),
    ]

    for r, (label, value) in enumerate(rows, start=1):
        lbl = ws.cell(row=r, column=1, value=label)
        val = ws.cell(row=r, column=2, value=value)
        lbl.font = _HEADER_FONT
        if isinstance(value, float):
            val.number_format = _NUM_FORMAT
        _set_border(lbl)
        _set_border(val)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22


# ---------------------------------------------------------------------------
# Utilidades de formato
# ---------------------------------------------------------------------------


def _write_header_row(ws, headers: list[str], row: int) -> None:
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        _set_border(cell)


def _set_border(cell) -> None:
    cell.border = _THIN_BORDER


def _autofit_columns(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(
            max_len + 2, 12
        )


def _all_spaces(building: Building) -> list[Space]:
    return [s for storey in building.storeys for s in storey.spaces]


def _residential_spaces(building: Building) -> list[Space]:
    return [space for space in _all_spaces(building) if not isinstance(space, ParkingSpace)]


def _parking_spaces(building: Building) -> list[ParkingSpace]:
    return [space for space in _all_spaces(building) if isinstance(space, ParkingSpace)]
